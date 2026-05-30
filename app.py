import base64
import io
import os
import uuid
from datetime import date, datetime

import cv2
import numpy as np
from dotenv import load_dotenv
from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import LoginManager, UserMixin, current_user, login_required, login_user, logout_user
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

load_dotenv()

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DEFAULT_UPLOAD_DIR = os.path.join(BASE_DIR, "static", "uploads", "teachers")
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", DEFAULT_UPLOAD_DIR)
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "change-this-secret-key")

raw_db_url = os.environ.get("DATABASE_URL", "sqlite:///instance/school_face_attendance.db")
if raw_db_url.startswith("postgres://"):
    raw_db_url = raw_db_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = raw_db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("MAX_CONTENT_LENGTH_MB", "30")) * 1024 * 1024

ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "webp"}
FACE_CONFIDENCE_THRESHOLD = int(os.environ.get("FACE_CONFIDENCE_THRESHOLD", "45"))
FACE_MARGIN_THRESHOLD = int(os.environ.get("FACE_MARGIN_THRESHOLD", "12"))
FACE_REQUIRED_MATCHES = int(os.environ.get("FACE_REQUIRED_MATCHES", "3"))
FACE_DETECTION_MIN_SIZE = int(os.environ.get("FACE_DETECTION_MIN_SIZE", "40"))
FACE_SKIP_QUALITY_CHECK = os.environ.get("FACE_SKIP_QUALITY_CHECK", "1") == "1"
FACE_ALLOW_CENTER_FALLBACK = os.environ.get("FACE_ALLOW_CENTER_FALLBACK", "0") == "1"



db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), default="admin", nullable=False)


class Teacher(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(160), nullable=False, unique=True)
    department = db.Column(db.String(160), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    images = db.relationship("TeacherImage", backref="teacher", cascade="all, delete-orphan")
    records = db.relationship("AttendanceRecord", backref="teacher", cascade="all, delete-orphan")


class TeacherImage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    teacher_id = db.Column(db.Integer, db.ForeignKey("teacher.id"), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)


class AttendanceRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    teacher_id = db.Column(db.Integer, db.ForeignKey("teacher.id"), nullable=False)
    record_date = db.Column(db.Date, nullable=False, default=date.today)
    sign_in_time = db.Column(db.DateTime, nullable=True)
    sign_out_time = db.Column(db.DateTime, nullable=True)
    sign_in_confidence = db.Column(db.Float, nullable=True)
    sign_out_confidence = db.Column(db.Float, nullable=True)
    status_note = db.Column(db.String(255), default="")


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def teacher_folder(teacher_id):
    path = os.path.join(UPLOAD_DIR, str(teacher_id))
    os.makedirs(path, exist_ok=True)
    return path


def image_to_gray_array(image_bytes):
    image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    array = np.array(image)
    bgr = cv2.cvtColor(array, cv2.COLOR_RGB2BGR)
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    return gray


def detect_largest_face(gray):
    """
    Tolerant face detection for school laptops with low-resolution cameras.
    This version does NOT reject faces because of blur/brightness quality.
    It first tries normal frontal detection, then relaxed frontal/profile detection.
    """
    if gray is None or gray.size == 0:
        return None

    # Improve weak laptop webcam frames without rejecting them.
    gray = cv2.resize(gray, None, fx=1.25, fy=1.25, interpolation=cv2.INTER_LINEAR)
    equalized = cv2.equalizeHist(gray)

    cascade_names = [
        "haarcascade_frontalface_default.xml",
        "haarcascade_frontalface_alt.xml",
        "haarcascade_frontalface_alt2.xml",
        "haarcascade_profileface.xml",
    ]

    candidates = []
    min_size = max(30, FACE_DETECTION_MIN_SIZE)
    for cascade_name in cascade_names:
        cascade_path = cv2.data.haarcascades + cascade_name
        detector = cv2.CascadeClassifier(cascade_path)
        if detector.empty():
            continue

        # Try multiple settings. minNeighbors=3 is more tolerant for weak webcams.
        for img in (equalized, gray):
            faces = detector.detectMultiScale(
                img,
                scaleFactor=1.05,
                minNeighbors=3,
                minSize=(min_size, min_size),
                flags=cv2.CASCADE_SCALE_IMAGE,
            )
            candidates.extend(list(faces))

        # Profile detector sometimes needs horizontally flipped image.
        if cascade_name == "haarcascade_profileface.xml":
            flipped = cv2.flip(equalized, 1)
            faces = detector.detectMultiScale(
                flipped,
                scaleFactor=1.05,
                minNeighbors=3,
                minSize=(min_size, min_size),
                flags=cv2.CASCADE_SCALE_IMAGE,
            )
            # Convert flipped x back to original coordinate space.
            width = equalized.shape[1]
            for x, y, w, h in faces:
                candidates.append((width - x - w, y, w, h))

    if not candidates:
        if FACE_ALLOW_CENTER_FALLBACK:
            h, w = gray.shape[:2]
            side = int(min(w, h) * 0.70)
            x = max(0, (w - side) // 2)
            y = max(0, (h - side) // 2)
            crop = equalized[y:y + side, x:x + side]
            return cv2.resize(crop, (200, 200)) if crop.size else None
        return None

    x, y, w, h = max(candidates, key=lambda f: f[2] * f[3])

    # Add small padding around detected face to include full facial features.
    pad = int(0.15 * max(w, h))
    x1 = max(0, x - pad)
    y1 = max(0, y - pad)
    x2 = min(equalized.shape[1], x + w + pad)
    y2 = min(equalized.shape[0], y + h + pad)
    face = equalized[y1:y2, x1:x2]
    if face.size == 0:
        return None
    return cv2.resize(face, (200, 200))


def decode_camera_image(data_url):
    if not data_url or "," not in data_url:
        raise ValueError("Camera image was not received correctly.")
    encoded = data_url.split(",", 1)[1]
    return base64.b64decode(encoded)


def face_quality_ok(face):
    """Basic quality gate to reduce false matches from blurry/dark camera frames."""
    blur_score = cv2.Laplacian(face, cv2.CV_64F).var()
    brightness = float(np.mean(face))
    if blur_score < 45:
        return False, "Face image is too blurry. Please hold still and improve focus."
    if brightness < 45:
        return False, "Face image is too dark. Please improve lighting."
    if brightness > 220:
        return False, "Face image is overexposed. Please reduce strong light."
    return True, "OK"


def collect_teacher_faces():
    teacher_faces = []
    teachers = Teacher.query.order_by(Teacher.name.asc()).all()
    for teacher in teachers:
        faces = []
        for img in teacher.images:
            path = os.path.join(teacher_folder(teacher.id), img.filename)
            if not os.path.exists(path):
                continue
            try:
                with open(path, "rb") as f:
                    gray = image_to_gray_array(f.read())
                face = detect_largest_face(gray)
                if face is not None:
                    faces.append(face)
            except Exception:
                continue
        if faces:
            teacher_faces.append((teacher, faces))
    return teacher_faces


def recognize_single_frame(image_bytes):
    captured_gray = image_to_gray_array(image_bytes)
    captured_face = detect_largest_face(captured_gray)
    if captured_face is None:
        return None, None, "No face was detected. Please make sure your face is inside the camera frame and try again."

    if not FACE_SKIP_QUALITY_CHECK:
        quality_ok, quality_message = face_quality_ok(captured_face)
        if not quality_ok:
            return None, None, quality_message

    teacher_faces = collect_teacher_faces()
    if not teacher_faces:
        return None, None, "No trained teacher images found. Admin must upload teacher photos first."

    scores = []
    for teacher, faces in teacher_faces:
        if len(faces) < 2:
            # one image is not enough for reliable recognition; still test it but with caution
            pass
        recognizer = cv2.face.LBPHFaceRecognizer_create(radius=1, neighbors=8, grid_x=8, grid_y=8)
        recognizer.train(faces, np.zeros(len(faces), dtype=np.int32))
        _, confidence = recognizer.predict(captured_face)
        scores.append((float(confidence), teacher))

    if not scores:
        return None, None, "No usable teacher face images found."

    scores.sort(key=lambda item: item[0])
    best_confidence, best_teacher = scores[0]
    second_confidence = scores[1][0] if len(scores) > 1 else 999.0
    margin = second_confidence - best_confidence

    if best_confidence > FACE_CONFIDENCE_THRESHOLD:
        return None, best_confidence, "Face not confidently matched. Please try again or ask admin to upload clearer photos."

    if len(scores) > 1 and margin < FACE_MARGIN_THRESHOLD:
        return None, best_confidence, "Face is too similar to another saved teacher photo. Match rejected for safety. Please try again with better lighting."

    return best_teacher, best_confidence, "Matched successfully."


def recognize_teacher(image_bytes):
    return recognize_single_frame(image_bytes)


def recognize_teacher_from_frames(image_list):
    successful = []
    last_message = "No face matched."
    last_confidence = None

    for data_url in image_list:
        try:
            teacher, confidence, message = recognize_single_frame(decode_camera_image(data_url))
            last_message = message
            last_confidence = confidence
            if teacher:
                successful.append((teacher.id, teacher, confidence))
        except Exception as exc:
            last_message = f"Processing error: {exc}"

    if not successful:
        return None, last_confidence, last_message

    counts = {}
    teachers = {}
    confidences = {}
    for teacher_id, teacher, confidence in successful:
        counts[teacher_id] = counts.get(teacher_id, 0) + 1
        teachers[teacher_id] = teacher
        confidences.setdefault(teacher_id, []).append(confidence)

    best_teacher_id = max(counts, key=counts.get)
    required = max(1, FACE_REQUIRED_MATCHES)
    if counts[best_teacher_id] < required:
        return None, None, f"Verification rejected for safety. Needed {required} matching frames, got {counts[best_teacher_id]}. Please try again."

    avg_confidence = sum(confidences[best_teacher_id]) / len(confidences[best_teacher_id])
    return teachers[best_teacher_id], float(avg_confidence), "Matched successfully across multiple frames."


def today_record_for(teacher):
    today = date.today()
    record = AttendanceRecord.query.filter_by(teacher_id=teacher.id, record_date=today).first()
    if not record:
        record = AttendanceRecord(teacher_id=teacher.id, record_date=today)
        db.session.add(record)
        db.session.flush()
    return record


def admin_required():
    return current_user.is_authenticated and current_user.role == "admin"


@app.before_request
def initialize_database():
    if not getattr(app, "_database_initialized", False):
        db.create_all()
        if not User.query.filter_by(username="admin").first():
            user = User(username="admin", password_hash=generate_password_hash("1234"), role="admin")
            db.session.add(user)
            db.session.commit()
        app._database_initialized = True


@app.route("/")
def index():
    teacher_count = Teacher.query.count()
    return render_template("index.html", teacher_count=teacher_count)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for("admin_dashboard"))
        flash("Invalid username or password.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("index"))


@app.route("/admin")
@login_required
def admin_dashboard():
    if not admin_required():
        return redirect(url_for("index"))
    teachers = Teacher.query.order_by(Teacher.name.asc()).all()
    today = date.today()
    records_today = AttendanceRecord.query.filter_by(record_date=today).all()
    return render_template("admin.html", teachers=teachers, records_today=records_today)


@app.route("/admin/teachers", methods=["POST"])
@login_required
def add_teacher():
    if not admin_required():
        return redirect(url_for("index"))
    name = request.form.get("name", "").strip()
    department = request.form.get("department", "").strip()
    files = request.files.getlist("photos")
    if not name:
        flash("Teacher name is required.", "danger")
        return redirect(url_for("admin_dashboard"))
    teacher = Teacher.query.filter_by(name=name).first()
    if not teacher:
        teacher = Teacher(name=name, department=department)
        db.session.add(teacher)
        db.session.flush()
    else:
        teacher.department = department or teacher.department
    saved = 0
    for file in files:
        if file and allowed_file(file.filename):
            ext = secure_filename(file.filename).rsplit(".", 1)[1].lower()
            filename = f"{uuid.uuid4().hex}.{ext}"
            path = os.path.join(teacher_folder(teacher.id), filename)
            raw = file.read()
            face = detect_largest_face(image_to_gray_array(raw))
            if face is None:
                continue
            with open(path, "wb") as f:
                f.write(raw)
            db.session.add(TeacherImage(teacher_id=teacher.id, filename=filename))
            saved += 1
    db.session.commit()
    flash(f"Teacher saved. {saved} valid face image(s) uploaded.", "success" if saved else "warning")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/teacher/<int:teacher_id>/delete", methods=["POST"])
@login_required
def delete_teacher(teacher_id):
    if not admin_required():
        return redirect(url_for("index"))
    teacher = db.session.get(Teacher, teacher_id)
    if teacher:
        folder = teacher_folder(teacher.id)
        db.session.delete(teacher)
        db.session.commit()
        try:
            for filename in os.listdir(folder):
                os.remove(os.path.join(folder, filename))
            os.rmdir(folder)
        except Exception:
            pass
        flash("Teacher deleted.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/attendance/<action>")
def attendance_page(action):
    if action not in {"in", "out"}:
        return redirect(url_for("index"))
    return render_template("attendance.html", action=action)


@app.route("/api/attendance", methods=["POST"])
def api_attendance():
    payload = request.get_json(silent=True) or {}
    action = payload.get("action")
    if action not in {"in", "out"}:
        return jsonify({"ok": False, "message": "Invalid attendance action."}), 400
    try:
        images = payload.get("images")
        if isinstance(images, list) and images:
            teacher, confidence, message = recognize_teacher_from_frames(images)
        else:
            image_bytes = decode_camera_image(payload.get("image"))
            teacher, confidence, message = recognize_teacher(image_bytes)
        if not teacher:
            return jsonify({"ok": False, "message": message, "confidence": confidence})
        record = today_record_for(teacher)
        now = datetime.now()
        if action == "in":
            if record.sign_in_time:
                return jsonify({"ok": False, "message": f"{teacher.name} already signed in today."})
            record.sign_in_time = now
            record.sign_in_confidence = confidence
            text = f"Welcome {teacher.name}. Sign in recorded at {now.strftime('%H:%M:%S')}."
        else:
            if not record.sign_in_time:
                return jsonify({"ok": False, "message": f"{teacher.name} must sign in before signing out."})
            if record.sign_out_time:
                return jsonify({"ok": False, "message": f"{teacher.name} already signed out today."})
            record.sign_out_time = now
            record.sign_out_confidence = confidence
            text = f"Goodbye {teacher.name}. Sign out recorded at {now.strftime('%H:%M:%S')}."
        db.session.commit()
        return jsonify({"ok": True, "message": text, "teacher": teacher.name, "confidence": confidence})
    except Exception as exc:
        return jsonify({"ok": False, "message": f"Processing error: {exc}"}), 500


def parse_date_or_none(value):
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def date_range(start, end):
    from datetime import timedelta
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def build_report_rows(teacher_id=None, start_date="", end_date="", status="all"):
    start = parse_date_or_none(start_date)
    end = parse_date_or_none(end_date)
    if status == "not_signed_in":
        if not start and not end:
            start = end = date.today()
        elif start and not end:
            end = start
        elif end and not start:
            start = end
        teachers_query = Teacher.query.order_by(Teacher.name.asc())
        if teacher_id:
            teachers_query = teachers_query.filter(Teacher.id == teacher_id)
        rows = []
        for day in date_range(start, end):
            for teacher in teachers_query.all():
                record = AttendanceRecord.query.filter_by(teacher_id=teacher.id, record_date=day).first()
                if record is None or record.sign_in_time is None:
                    rows.append({
                        "date": day,
                        "teacher": teacher,
                        "sign_in_time": record.sign_in_time if record else None,
                        "sign_out_time": record.sign_out_time if record else None,
                        "status": "Not signed in",
                        "sign_in_confidence": record.sign_in_confidence if record else None,
                        "sign_out_confidence": record.sign_out_confidence if record else None,
                    })
        return rows

    query = AttendanceRecord.query.join(Teacher)
    if teacher_id:
        query = query.filter(AttendanceRecord.teacher_id == teacher_id)
    if start:
        query = query.filter(AttendanceRecord.record_date >= start)
    if end:
        query = query.filter(AttendanceRecord.record_date <= end)
    if status == "not_signed_out":
        query = query.filter(AttendanceRecord.sign_in_time.isnot(None), AttendanceRecord.sign_out_time.is_(None))
    elif status == "complete":
        query = query.filter(AttendanceRecord.sign_in_time.isnot(None), AttendanceRecord.sign_out_time.isnot(None))
    records = query.order_by(AttendanceRecord.record_date.desc(), Teacher.name.asc()).all()
    rows = []
    for record in records:
        if not record.sign_in_time:
            row_status = "Not signed in"
        elif not record.sign_out_time:
            row_status = "Not signed out"
        else:
            row_status = "Complete"
        rows.append({
            "date": record.record_date,
            "teacher": record.teacher,
            "sign_in_time": record.sign_in_time,
            "sign_out_time": record.sign_out_time,
            "status": row_status,
            "sign_in_confidence": record.sign_in_confidence,
            "sign_out_confidence": record.sign_out_confidence,
        })
    return rows


@app.route("/admin/reports")
@login_required
def reports():
    if not admin_required():
        return redirect(url_for("index"))
    teacher_id = request.args.get("teacher_id", type=int)
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    status = request.args.get("status", "all")
    rows = build_report_rows(teacher_id, start_date, end_date, status)
    teachers = Teacher.query.order_by(Teacher.name.asc()).all()
    return render_template("reports.html", rows=rows, teachers=teachers, filters=request.args)


@app.route("/admin/reports/export")
@login_required
def export_report():
    if not admin_required():
        return redirect(url_for("index"))
    teacher_id = request.args.get("teacher_id", type=int)
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    status = request.args.get("status", "all")
    rows = build_report_rows(teacher_id, start_date, end_date, status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    ws.merge_cells("A1:H1")
    ws["A1"] = "School Face Recognition Attendance Report"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["Date", "Teacher", "Department", "Sign In", "Sign Out", "Status", "Sign In Confidence", "Sign Out Confidence"]
    ws.append([])
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(border_style="thin", color="D9D9D9")
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([
            row["date"].strftime("%Y-%m-%d"),
            row["teacher"].name,
            row["teacher"].department or "-",
            row["sign_in_time"].strftime("%H:%M:%S") if row["sign_in_time"] else "-",
            row["sign_out_time"].strftime("%H:%M:%S") if row["sign_out_time"] else "-",
            row["status"],
            round(row["sign_in_confidence"], 2) if row["sign_in_confidence"] is not None else "-",
            round(row["sign_out_confidence"], 2) if row["sign_out_confidence"] is not None else "-",
        ])
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center")
    widths = [15, 28, 22, 14, 14, 18, 20, 20]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A4"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
